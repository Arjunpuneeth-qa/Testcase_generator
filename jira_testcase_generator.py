#!/usr/bin/env python3
"""
JIRA Test Case Generator Agent
Fetches JIRA tickets and generates test cases in Excel format
"""

import os
import sys
import requests
import json
from dotenv import load_dotenv

load_dotenv()
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import base64

class JiraTestCaseGenerator:
    def __init__(self, jira_url, email, api_token):
        self.jira_url = jira_url.rstrip('/')
        self.email = email
        self.api_token = api_token
        self.session = self._create_session()
        self.output_dir = os.path.join(os.path.dirname(__file__), 'GA_testcases')
        os.makedirs(self.output_dir, exist_ok=True)

    def _create_session(self):
        """Create authenticated session"""
        session = requests.Session()
        credentials = f"{self.email}:{self.api_token}"
        encoded = base64.b64encode(credentials.encode()).decode()
        session.headers.update({
            'Authorization': f'Basic {encoded}',
            'Content-Type': 'application/json'
        })
        return session

    def fetch_ticket(self, ticket_key):
        """Fetch ticket from JIRA API"""
        try:
            url = f"{self.jira_url}/rest/api/3/issues/{ticket_key}"
            response = self.session.get(url)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"Error fetching ticket {ticket_key}: {e}")
            sys.exit(1)

    def parse_ticket_to_testcases(self, ticket):
        """Parse JIRA ticket and generate test cases"""
        fields = ticket.get('fields', {})

        ticket_key = ticket.get('key', 'UNKNOWN')
        summary = fields.get('summary', 'No Summary')
        description = fields.get('description', '')
        ticket_type = fields.get('issuetype', {}).get('name', 'Task')

        # Parse description for test steps or acceptance criteria
        test_cases = []

        # Default test case structure
        test_case = {
            'Test Case ID': f"{ticket_key}_TC001",
            'Test Case Name': f"{ticket_key} - {summary}",
            'Description': description if isinstance(description, str) else (description.get('plain_text', '') if isinstance(description, dict) else ''),
            'Preconditions': 'Ticket created and assigned',
            'Test Steps': '1. Verify ticket requirements',
            'Expected Result': 'Requirements met',
            'Test Status': 'Not Started',
            'Created Date': datetime.now().strftime('%Y-%m-%d'),
            'Updated Date': datetime.now().strftime('%Y-%m-%d'),
            'Ticket Type': ticket_type
        }

        test_cases.append(test_case)

        # If description contains numbered items, parse them as test steps
        if test_case['Description']:
            lines = test_case['Description'].split('\n')
            steps = []
            for line in lines:
                if line.strip() and (line[0].isdigit() or line.strip().startswith('-')):
                    steps.append(line.strip())
            if steps:
                test_case['Test Steps'] = '\n'.join(steps[:10])

        return test_cases, ticket_key, summary

    def create_excel_file(self, test_cases, ticket_key, summary):
        """Create Excel workbook with test cases"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'Test Case ID',
            'Test Case Name',
            'Description',
            'Preconditions',
            'Test Steps',
            'Expected Result',
            'Test Status',
            'Created Date',
            'Updated Date',
            'Ticket Type'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Write test cases
        for row_idx, test_case in enumerate(test_cases, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = test_case.get(header, '')
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = border

        # Adjust column widths
        column_widths = {
            'A': 15,  # Test Case ID
            'B': 30,  # Test Case Name
            'C': 25,  # Description
            'D': 20,  # Preconditions
            'E': 25,  # Test Steps
            'F': 25,  # Expected Result
            'G': 15,  # Test Status
            'H': 12,  # Created Date
            'I': 12,  # Updated Date
            'J': 15   # Ticket Type
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Set row heights
        ws.row_dimensions[1].height = 30
        for row in range(2, len(test_cases) + 2):
            ws.row_dimensions[row].height = 50

        # Create filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{ticket_key}_{summary.replace(' ', '_')[:30]}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        # Save workbook
        wb.save(filepath)
        return filepath

    def generate(self, ticket_key):
        """Main method to generate test cases from ticket"""
        print(f"\n{'='*60}")
        print(f"JIRA Test Case Generator")
        print(f"{'='*60}")
        print(f"\nFetching ticket: {ticket_key}...")

        # Fetch ticket
        ticket = self.fetch_ticket(ticket_key)
        print(f"✓ Ticket fetched successfully")

        # Parse ticket
        test_cases, ticket_key_from_response, summary = self.parse_ticket_to_testcases(ticket)
        print(f"✓ Test cases parsed: {len(test_cases)} test case(s) created")

        # Create Excel
        filepath = self.create_excel_file(test_cases, ticket_key_from_response, summary)
        print(f"✓ Excel file created")

        print(f"\n{'='*60}")
        print(f"Success! Test cases generated.")
        print(f"Output File: {filepath}")
        print(f"Location: {self.output_dir}")
        print(f"{'='*60}\n")

        return filepath


def main():
    """Main entry point"""
    # Configuration
    JIRA_URL = "https://bounteous.jira.com"
    EMAIL = os.getenv("JIRA_EMAIL")
    API_TOKEN = os.getenv("JIRA_API_TOKEN")

    # Get ticket key from command line
    if len(sys.argv) < 2:
        print("Usage: python jira_testcase_generator.py <TICKET_KEY>")
        print("Example: python jira_testcase_generator.py GAAM-618")
        sys.exit(1)

    ticket_key = sys.argv[1].upper()

    # Generate test cases
    generator = JiraTestCaseGenerator(JIRA_URL, EMAIL, API_TOKEN)
    generator.generate(ticket_key)


if __name__ == "__main__":
    main()
